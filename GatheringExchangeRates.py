
# 作者信息
# =========================================================================

# 作者：写代码的会计（同名微信公众号与 bilibili 视频号）
# 邮箱：iCodeAcc@qq.com
# github: https://github.com/1CodeAcc

# =========================================================================


# 免责声明
# =========================================================================

# 程序代码仅供学习交流使用，严禁用于商业或非法用途。
# 作者不承担因此代码被乱用产生的一切后果与法律责任。

# =========================================================================


# 本程序代码使用的命名规范
# =========================================================================

# 变量：有两种形式。
#       1) 直接为小写单词，如 cell、browser。
#       2) 匈牙利命名法，如 nThisIsAnInteger、lstThisIsAList、wsThisIsAWorkSheetObject。
# 函数：大驼峰命名法，如 DoSomething()。

# =========================================================================


# OS : Windows 10 1909
# IDE: Visual Studio Code


# WebDriver
# =========================================================================

# 使用 selenium 操作浏览器，需要下载该浏览器对应的 WebDriver；
# 将 WebDriver 程序放入 PATH 环境变量，即可在代码中直接调用。
# Firefox: https://github.com/mozilla/geckodriver/releases
# IE     : https://selenium-release.storage.googleapis.com/index.html
# Safari : https://developer.apple.com/documentation/webkit/about_webdriver_for_safari
# Opera  : https://github.com/operasoftware/operachromiumdriver/releases
# Chrome : https://sites.google.com/a/chromium.org/chromedriver/
# Edge   : https://developer.microsoft.com/en-us/microsoft-edge/tools/webdriver/

# =========================================================================


# HTML元素基础概念
# =========================================================================

# 假定有如下 HTML 代码：
# <td width="110px;" align="left">
#     <select name="pjname" id="pjname">
# 	      <option value="0">选择货币</option>
# 	      <option value="欧元">欧元</option>
# 	      <option value="日元">日元</option>
#         <option value="卢布">卢布</option>
#     </select>
# </td>

# "td"、"select"、"option" 称为 "标签(tag)"，标签都是以 "<tag>...</tag>" 的形式成对出现的；
# 标签及其中的内容构成一个网页元素(element)。
# width="110px;"、name="pjname"、value="0" 等称为属性(attribute)，属性以 "属性名=属性值" 的形式定义。
# "选择货币"、"欧元"、"日元" 等是元素中的文本(text)。
# 对于 <option value="欧元">欧元</option> 及其后的两个元素，其 value 属性的值和元素中的文本是相同的。

# =========================================================================


# selenium 查找网页元素函数
# =========================================================================

# find_element(s)_by_id()                   通过 id 查找
# find_element(s)_by_name()                 通过 name 查找
# find_element(s)_by_xpath()                通过 xpath 查找
# find_element(s)_by_tag_name()             通过标签查找
# find_element(s)_by_link_text()            通过包含连接的文本查找
# find_element(s)_by_class_name()           通过 class 查找
# find_element(s)_by_css_selector()         通过 CSS 选择器查找
# find_element(s)_by_partial_link_text()    通过包含连接的文本关键字查找

# 在编写代码前，应先使用浏览器的“检查元素”功能查看需要定位的元素的代码，再选择要使用的函数。

# =========================================================================


# 主要相关程序及软件包版本
# =========================================================================

# 本程序代码在如下版本程序及软件包下成功运行：
# selenium    3.141.0
# Firefox     78.3.1esr
# geckodirver 0.27.0
# openpyxl    3.0.5

# 若程序运行时总是提示找不到元素或发生其他与元素定位有关的异常，
# 则可能是因为 selenium 与 WebDriver 版本不匹配，可尝试更换二者版本。

# =========================================================================


# 步骤
# =========================================================================

# 1. 获取当天日期；新键表格对象、表格属性设置。
# 2. 爬取汇率信息并填写至表格、设置单元格格式。
# 3. 保存结果。

# =========================================================================


# =========================================================================

import os
import colorama
import datetime
import openpyxl
from selenium import webdriver
from colorama import Fore, Style
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, colors

# =========================================================================


# =========================================================================

# 在控制台打印带有颜色的内容
# *args 表示可变参数，是一个元组；args 是约定俗成的写法，也可使用其他变量名。
def ColorPrint(*args):
    colorama.init()
    sText = ""
    for arg in args:
        sText += arg
    print(sText)
    print(Style.RESET_ALL, end="")
    colorama.deinit()

# =========================================================================


# 单元格样式
# =========================================================================

# 边线样式
sideThin = Side(border_style = "thin", color = colors.BLACK)
# 边框样式
borderThin = Border(
                     left   = sideThin,    # 左边线
                     right  = sideThin,    # 右边线
                     top    = sideThin,    # 上边线
                     bottom = sideThin     # 下边线
                   )

# 字体和字号
# 使用等宽字体会让结果看起来更整齐。
# 宋体，11号。
fontSong     = Font(name = "宋体", size = 11)
# Consolas，11号。
fontConsolas = Font(name = "Consolas", size = 11)

# 顶端标题行填充颜色
pfTitle  = PatternFill(
                        fill_type = "solid",    # 纯色填充
                        fgColor   = "C0C0C0"    # 前景色
                                                # openpyxl 指定前景色为单元格填充颜色
                      )

# 单元格对齐方式
# 水平靠左，垂直居中。
alignHLeftVCenter   = Alignment(
                                 horizontal = "left",     # 水平对齐方式
                                 vertical   = "center"    # 垂直对齐方式
                                 # wrap_text = True       # 自动换行
                               )
# 水平居中，垂直居中。
alignHCenterVCenter = Alignment(
                                 horizontal = "center",
                                 vertical   = "center"
                               )
# 水平靠右，垂直居中。
alignHRightVCenter  = Alignment(
                                 horizontal = "right",
                                 vertical   = "center"
                               )

# 设置行高
def SetRowsDimensions(wsObj, nStartRow, nEndRow, nRowHeight):
    while nStartRow <= nEndRow:
        wsObj.row_dimensions[nStartRow].height = nRowHeight
        nStartRow += 1
    return

# 设置列宽
def SetColumnsDimensions(wsObj, lstColumns, lstColumnsWidth):
    # zip 函数可平行遍历多个可迭代对象；
    # 当多个可迭代对象不等长时，最短的对象遍历完即停止遍历。
    for sColumnIndex, nColumnWidth in zip(lstColumns, lstColumnsWidth):
        wsObj.column_dimensions[sColumnIndex].width = nColumnWidth
    return

# 设置标题行单元格格式
def SetTitleRowCellsFormat(wsObj, sCellsArea, border, font, fill, align):
    # 2*2或以上矩阵单元格区域为一个二维元组；
    # 二维元组的各个元素为每行 <Cell> 对象类型组成的一维元组。
    for tupRows in wsObj[sCellsArea]:
        for cell in tupRows:
            cell.border    = border
            cell.font      = font
            cell.fill      = fill
            cell.alignment = align
    return

# 设置列的单元格格式
def SetColumnCellsFormat(wsObj, sColumnIndex, border, font, align, tupIgnoreRows):
    # 某一行（或列）单元格区域为该行（或列）存在数据的单元格 <Cell> 对象类型组成的一维元组。
    for cell in wsObj[sColumnIndex]:
        # 不进行设置的行
        if cell.row in tupIgnoreRows:
            continue
        cell.border    = border
        cell.font      = font
        cell.alignment = align
    return

# 设置数字列的数字格式
def SetColumnNumbersFormat(wsObj, sColumnIndex, sFormatString):
    for cell in wsObj[sColumnIndex]:
        # 若单元格数据不为数值则不进行设置
        if not isinstance(cell.value, int) and not isinstance(cell.value, float):
            continue
        cell.number_format = sFormatString
    return

# =========================================================================


# 日期、表名、文件名。
# =========================================================================

dtToday = datetime.date.today()
# %y 代表两位数年份(00-99)，%Y 代表四位数年份(0000-9999)。
# %m 代表月份(01-12)，%M 代表分钟(00-59)。
# %d 代表天数(01-31)，无 %D 形式。
sTodayForTitle   = dtToday.strftime("%Y%m%d")
sTodayForSearch  = dtToday.strftime("%Y-%m-%d")
# 工作表及工作簿保存名称
sSaveName = "FX-" + sTodayForTitle
sSaveFilePath = os.getcwd() + "\\" + sSaveName + ".xlsx"

wbFX = openpyxl.Workbook()    # 创建工作簿对象（但并不新建文件）
wsFX = wbFX.active            # 进入活动工作表（默认新键后只有一个工作表，即为活动工作表）
wsFX.title = sSaveName        # 修改工作表表名

# =========================================================================


# 获取币种代码
# =========================================================================

# 在工作簿末尾插入新表
wsCodes = wbFX.create_sheet("Codes")
# 顶端标题行
# append 方法将从最后一个被修改过的行（包括写入数据、设置单元格格式、设置行高列宽等）
# 的下一行的最左侧单元格开始，依次将列表中的数据写入该行的各个单元格。
# 每调用一次 append 方法，按上述规则在新的一行写入数据。
wsCodes.append(["币种", "代码"])

# 浏览器无头模式（即不显示浏览器窗口）
profile = webdriver.FirefoxOptions()
profile.add_argument("-headless")
browser = webdriver.Firefox(options = profile)

# 访问页面
# 资料来源：站长之家。
browser.get("http://www.webmasterhome.cn/huilv/huobidaima.asp")
# 隐式等待元素加载：
# 指定时间内元素未加载完毕，则不再等待，代码继续执行；
# 指定时间内元素加载完毕，在元素加载完毕后代码继续执行。
browser.implicitly_wait(5)
# 访问页面并等待页面元素加载完毕后，即可进行网页元素的查找定位。
# 所有币种信息对应元素
elemCurrenciesList = browser.find_elements_by_tag_name("tr")
print("")
print("获取币种代码...")
# 循环时跳过第一个（标题行对应元素）
for elemCurrency in elemCurrenciesList[1:]:
    # 如下被注释代码得不到正确结果（均为 "人民币:CNY"），具体原因未知。
    # sCurrency = elemCurrency.find_element_by_xpath("//td/a").get_attribute("title")
    # sCode     = elemCurrency.find_element_by_xpath("//td/span").get_attribute("class")
    sCurrency = elemCurrency.find_element_by_tag_name("a").get_attribute("title")
    sCode     = elemCurrency.find_element_by_tag_name("span").get_attribute("class")
    # 名称与中行不一致的，按中行名称修改。
    if sCurrency == "澳元":
        sCurrency = "澳大利亚元"
    elif sCurrency == "加元":
        sCurrency = "加拿大元"
    elif sCurrency == "泰铢":
        sCurrency = "泰国铢"
    elif sCurrency == "韩元":
        sCurrency = "韩国元"
    elif sCurrency == "俄罗斯卢布":
        sCurrency = "卢布"
    elif sCurrency == "印度尼西亚卢比":
        sCurrency = "印尼卢比"
    elif sCurrency == "巴西雷亚尔":
        sCurrency = "巴西里亚尔"
    else:
        pass
    wsCodes.append([sCurrency, sCode])
    print("%s : %s" % (sCode, sCurrency))
# 添加中行有、但站长之家没有的币种。
lstExtraAddedCurrencies = [
                            ["西班牙比塞塔", "ESP"],
                            ["比利时法郎", "BEF"],
                            ["芬兰马克", "FIM"]
                          ]
for lstCurrency in lstExtraAddedCurrencies:
    wsCodes.append(lstCurrency)
    print("%s : %s" % (lstCurrency[1], lstCurrency[0]))
print("完成")

# 设置单元格格式
wsCodes.freeze_panes = "A2"    # 冻结首行
SetRowsDimensions(wsCodes, 1, wsCodes.max_row, 19)
SetColumnsDimensions(wsCodes, ["A", "B"], [23, 8])
SetTitleRowCellsFormat(wsCodes, "A1:B1", borderThin, fontSong, pfTitle, alignHCenterVCenter)
SetColumnCellsFormat(wsCodes, "A", borderThin, fontSong,     alignHLeftVCenter,   (1,))
SetColumnCellsFormat(wsCodes, "B", borderThin, fontConsolas, alignHCenterVCenter, (1,))

# =========================================================================


# 输出内容对齐方法
# =========================================================================

# str.format() 方法中的替换字段(replacement field) "{}" 的常用格式
# （中括号内为可选内容，竖线相隔内容为不同选项，省略号表示参数个数可变）：
# "{[para_index|para_name][:][filled_char][<|^|>][output_len]}".format([para_name=]const|var, ...)
# para_index  : 参数索引，从 0 开始。
# para_name   : 参数名。
# filled_char : 对齐填充字符，默认为 ASCII 空格。
# <|^|>       : 分别表示居左、句中、居右对齐。
# output_len  : 输出宽度；对于数字类型，可使用 "d"、"f" 等格式化字符。
# const       : 常量。
# var         : 已定义的变量名。
# 包含中文的文本对齐需使用 Unicode 空格，可使用 chr(12288) 表示。

# =========================================================================


# 获取中行汇率
# =========================================================================

# 顶端标题行
wsFX.append(["币种", "代码", "汇率"])

browser.get("https://www.boc.cn/sourcedb/whpj/")
browser.implicitly_wait(5)
# 所有币种名称对应元素
elemSearchCurrencies = browser.find_elements_by_tag_name("option")
# 提取所有币种的名称
lstSearchCurrencies  = []
# 跳过标题行对应元素
for elemSearchCurrency in elemSearchCurrencies[1:]:
    lstSearchCurrencies.append(elemSearchCurrency.get_attribute("value"))
# max 函数除可求多个参数中的最大值，还可求可迭代对象中的最大值。
# 参数1：必须为非空可迭代对象。
# 参数2：可选，指定取最大值的计算尺度函数。
# 参数3：可选，当取不到最大值时的返回值。
nMaxLenCurrencyName = len(max(lstSearchCurrencies, key = len, default = None))
print("")
print("获取汇率...")
# 构造查询 URL：请求网址(Requst URL) + 参数(Parameters)。
# 请求网址和参数均可在浏览器开发者工具中获得。
# 开发者工具 —— 网络 —— 选择第一条 POST/GET 方法 —— HTML —— 消息头/参数。
sRequstUrl = "https://srh.bankofchina.com/search/whpj/search_cn.jsp"
for sSearchCurrency in lstSearchCurrencies:
    sParameters = "?erectDate={0}&nothing={0}&pjname={1}".format(sTodayForSearch, sSearchCurrency)
    sSearchUrl  = sRequstUrl + sParameters
    browser.get(sSearchUrl)
    browser.implicitly_wait(5)
    try:
        # 中行折算价第一行记录
        # xpath 定位元素可在路径中使用索引，从而简化多层查找；
        # xpath 中的元素索引从 1 开始。
        browser.find_element_by_xpath("//tr[@class='odd']/td[6]")
    except:
        fFX = sFX = "None"
    else:
        sFXTemp = browser.find_element_by_xpath("//tr[@class='odd']/td[6]").text
        fFX = float(sFXTemp) / 100
        sFX = str("{:.6f}".format(fFX))
    wsFX.append([sSearchCurrency, None, fFX])
    print("{search_currency:{filled_char}>{output_len}} : {fx}".format(search_currency = sSearchCurrency, 
                                                                       fx = sFX,
                                                                       filled_char = chr(12288), 
                                                                       output_len  = nMaxLenCurrencyName))
print("完成")

# 使用 VLOOKUP 从 Codes 表中获取币种代码
# 跳过标题行单元格
# 循环时跳过可迭代对象第一个元素的另一种方法
tupiterCurrencies = iter(wsFX["A"])
tupiterCodes      = iter(wsFX["B"])
next(tupiterCurrencies)
next(tupiterCodes)
for cellCurrency, cellCode in zip(tupiterCurrencies, tupiterCodes):
    cellCode.value = "=VLOOKUP({},{}!A:B,2,0)".format(cellCurrency.coordinate, wsCodes.title)

# 设置单元格格式
wsCodes.freeze_panes = "A2"
SetRowsDimensions(wsFX, 1, wsFX.max_row, 19)
SetColumnsDimensions(wsFX, ["A", "B", "C"], [23, 8, 13])
SetTitleRowCellsFormat(wsFX, "A1:C1", borderThin, fontSong, pfTitle, alignHCenterVCenter)
SetColumnCellsFormat(wsFX, "A", borderThin, fontSong,     alignHLeftVCenter,   (1,))
SetColumnCellsFormat(wsFX, "B", borderThin, fontConsolas, alignHCenterVCenter, (1,))
SetColumnCellsFormat(wsFX, "C", borderThin, fontConsolas, alignHRightVCenter,  (1,))
# 数字列数字格式：千位分隔符、负数显示负号、保留 6 位小数。
# 设置单元格格式的格式化字符串获取方法：
# 先在 Excel 的“设置单元格格式”中，设置好单元格格式（比如在“数字”项下设置数字的格式），
# 然后转到“自定义”项下，“类型”下面的字符串即为格式化字符串。
SetColumnNumbersFormat(wsFX, "C", "#,##0.000000_ ")

# =========================================================================


# 收尾
# =========================================================================

print("")
ColorPrint("保存文件: ", Style.BRIGHT + Fore.CYAN, "%s" % sSaveFilePath)
# 保存文件
# 保存将直接覆盖同名文件，不会有任何提示。
wbFX.save("%s.xlsx" % sSaveName)
print("")
print("全部完成")
print("")

while True:
    sOpenFileOrNot = input("是否打开文件(Y/N)? ")
    if sOpenFileOrNot == "Y" or sOpenFileOrNot == "y":
        # 命令行打开文件，直接输入文件路径即可。
        os.system('\"%s\"' % sSaveFilePath)
    elif sOpenFileOrNot == "N" or sOpenFileOrNot == "n":
        pass
    else:
        print("输入错误！")
        continue
    break

# =========================================================================